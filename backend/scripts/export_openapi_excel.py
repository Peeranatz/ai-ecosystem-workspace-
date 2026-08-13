import os
import sys
import json
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Ensure backend root is in sys.path
backend_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if backend_dir not in sys.path:
    sys.path.insert(0, backend_dir)

from app.main import app

def generate_openapi_snapshot():
    print("Extracting OpenAPI Schema from FastAPI application...")
    openapi_schema = app.openapi()
    
    paths = openapi_schema.get("paths", {})
    api_records = []
    
    for path, methods in paths.items():
        for method_name, details in methods.items():
            if method_name.upper() not in ["GET", "POST", "PUT", "PATCH", "DELETE"]:
                continue
                
            verb = method_name.upper()
            tags = details.get("tags", ["Uncategorized"])
            tag_name = tags[0] if tags else "Uncategorized"
            summary = details.get("summary", "")
            description = details.get("description", "")
            
            # Extract parameters
            params_list = []
            for p in details.get("parameters", []):
                p_name = p.get("name", "")
                p_in = p.get("in", "")
                p_req = "Required" if p.get("required", False) else "Optional"
                p_type = p.get("schema", {}).get("type", "string")
                params_list.append(f"{p_name} ({p_in}, {p_type}, {p_req})")
                
            if "requestBody" in details:
                params_list.append("request_body (application/json)")
                
            params_str = "\n".join(params_list) if params_list else "None"
            
            # Extract response status codes
            responses = details.get("responses", {})
            responses_str = ", ".join([str(code) for code in responses.keys()])
            
            api_records.append({
                "tag": tag_name,
                "path": path,
                "verb": verb,
                "summary": summary,
                "description": description.strip(),
                "parameters": params_str,
                "responses": responses_str
            })
            
    print(f"Total API Endpoints Extracted: {len(api_records)}")
    
    # 1. Export CSV
    csv_path = os.path.join(backend_dir, "api_list_snapshot.csv")
    with open(csv_path, mode="w", newline="", encoding="utf-8-sig") as csv_file:
        writer = csv.DictWriter(csv_file, fieldnames=["tag", "path", "verb", "summary", "description", "parameters", "responses"])
        writer.writeheader()
        writer.writerows(api_records)
    print(f"CSV Snapshot exported to: {csv_path}")

    # 2. Export Excel (.xlsx) with Professional Styling
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "API List Snapshot"
    
    # Title Block
    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = "AI Ecosystem Enterprise Web API - Snapshot List"
    title_cell.font = Font(name="Segoe UI", size=16, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40
    
    # Subtitle / Info Block
    ws.merge_cells("A2:G2")
    sub_cell = ws["A2"]
    sub_cell.value = "Generated automatically from FastAPI OpenAPI Schema | Clean Architecture Framework"
    sub_cell.font = Font(name="Segoe UI", size=10, italic=True, color="595959")
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20
    
    # Header Row
    headers = ["Domain / Tag", "API Endpoint Path", "HTTP Verb", "Summary", "Description", "Parameters", "Status Codes"]
    ws.append([]) # Row 3 blank
    ws.append(headers) # Row 4
    
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9")
    )
    
    ws.row_dimensions[4].height = 28
    for col_num in range(1, 8):
        c = ws.cell(row=4, column=col_num)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin_border

    # Verb Color Fills
    verb_fills = {
        "GET": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
        "POST": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
        "PUT": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
        "DELETE": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    }

    # Data Rows
    row_start = 5
    for idx, r in enumerate(api_records, start=row_start):
        ws.append([
            r["tag"],
            r["path"],
            r["verb"],
            r["summary"],
            r["description"],
            r["parameters"],
            r["responses"]
        ])
        
        ws.row_dimensions[idx].height = 24
        for col_num in range(1, 8):
            cell = ws.cell(row=idx, column=col_num)
            cell.font = Font(name="Segoe UI", size=10)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            
            if col_num == 3: # HTTP Verb Column
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(name="Segoe UI", size=10, bold=True)
                if r["verb"] in verb_fills:
                    cell.fill = verb_fills[r["verb"]]
                    
            if col_num in [1, 7]:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # Set Column Widths
    column_widths = {
        "A": 32, # Tag
        "B": 36, # Path
        "C": 14, # Verb
        "D": 32, # Summary
        "E": 40, # Description
        "F": 38, # Parameters
        "G": 16  # Responses
    }
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    excel_path = os.path.join(backend_dir, "api_list_snapshot.xlsx")
    wb.save(excel_path)
    print(f"Excel Snapshot exported successfully to: {excel_path}")

if __name__ == "__main__":
    generate_openapi_snapshot()
